
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.styles import Font
from django.http.response import HttpResponse
from apps.mantenimiento.models import Factura


# -------------------------------------------------------->
# Reporte Mantenimiento


class ReporteMantenimiento(TemplateView):
    def get(self, request, *args, **kwargs):
        facturas = Factura.objects.all()
        for fac in facturas:
            mantenimiento = fac.mantenimiento_set.all()
            if mantenimiento:
               comentarios = list(mantenimiento.values('comentario').values_list('comentario', flat=True))
               comentarios = ', '.join(comentarios)

               autorizado = list(mantenimiento.values('autorizado').values_list('autorizado', flat=True))
               autorizado = ', '.join(autorizado)

               tipoMant = list(mantenimiento.values('mant__nombre').values_list('mant__nombre', flat=True))
               tipoMant = ', '.join(str(tipoMant) for tipoMant in tipoMant)

               empresa = list(mantenimiento.values('empresa__nombre').values_list('empresa__nombre', flat=True))
               empresa = ', '.join(str(empresa) for empresa in empresa)

               realizado = list(mantenimiento.values('realizado').values_list('realizado', flat=True))
               realizado = ', '.join(realizado)

               fechaMant = list(mantenimiento.values('fecha').values_list('fecha', flat=True))
               fechaMant = ', '.join(str(fechaMant) for fechaMant in fechaMant)
            else:
                comentarios = 'No existe registro'
                autorizado = 'No existe registro'
                tipoMant = 'No existe registro'
                empresa = 'No existe registro'
                realizado = 'No existe registro'
                fechfechaMant = 'No existe registro'

            fac.comentarios = comentarios
            fac.autorizado = autorizado
            fac.tipoMant = tipoMant
            fac.empresa = empresa
            fac.realizado = realizado
            fac.fechaMant = fechaMant

        wb = Workbook()
        ws = wb.active

        thin = Side(border_style="thin", color="000000")  # Estilo de frontera, color
        border = Border(left=thin, right=thin, top=thin, bottom=thin)  # Posición fronteriza

        titulo = Font(size=14, bold=True, name='Microsoft Yahei', color="4d148c")
        titulo2 = Font(size=14, bold=True, name='Microsoft Yahei', color="ff6600")
        font = Font(size=8, bold=True, name='Microsoft Yahei', color="000000")

        fill = PatternFill(patternType="solid", start_color="ff6600")
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws['B2'] = 'Reporte De Mantenimiento'
        ws.merge_cells('B2:E2')
        ws['B2'].font = titulo

        ws['B3'] = 'Passus Velox'
        ws.merge_cells('B3:E3')
        ws['B3'].font = titulo2

        ws['B5'] = 'Factura '
        ws['B5'].font = font
        ws['B5'].alignment = align
        ws['B5'].border = border

        ws['C5'] = 'Proveedor '
        ws['C5'].font = font
        ws['C5'].alignment = align
        ws['C5'].border = border

        ws['D5'] = 'Total C$ '
        ws['D5'].font = font
        ws['D5'].alignment = align
        ws['D5'].border = border

        ws['E5'] = 'SubTotal '
        ws['E5'].font = font
        ws['E5'].alignment = align
        ws['E5'].border = border

        ws['F5'] = 'IVA '
        ws['F5'].font = font
        ws['F5'].alignment = align
        ws['F5'].border = border

        ws['G5'] = 'T/C '
        ws['G5'].font = font
        ws['G5'].alignment = align
        ws['G5'].border = border

        ws['H5'] = 'Total $'
        ws['H5'].font = font
        ws['H5'].alignment = align
        ws['H5'].border = border

        ws['I5'] = 'Detalle Factu6 '
        ws['I5'].font = font
        ws['I5'].alignment = align
        ws['I5'].border = border

        ws['J5'] = 'Fecha factura '
        ws['J5'].font = font
        ws['J5'].alignment = align
        ws['J5'].border = border

        ws['K5'] = 'Empresa '
        ws['K5'].font = font
        ws['K5'].alignment = align
        ws['K5'].border = border

        ws['L5'] = 'Autorizado '
        ws['L5'].font = font
        ws['L5'].alignment = align
        ws['L5'].border = border

        ws['M5'] = 'Tipo de mantenimiento '
        ws['M5'].font = font
        ws['M5'].alignment = align
        ws['M5'].border = border

        ws['N5'] = 'Comentarios del mantenimiento'
        ws['N5'].font = font
        ws['N5'].alignment = align
        ws['N5'].border = border

        ws['O5'] = 'Realizado '
        ws['O5'].font = font
        ws['O5'].alignment = align
        ws['O5'].border = border

        ws['P5'] = 'Fecha mantenimiento'
        ws['P5'].font = font
        ws['P5'].alignment = align
        ws['P5'].border = border

        cont = 6

        for fac in facturas:

            ws.cell(row=cont, column=2).value = fac.numFac
            ws.cell(row=cont, column=2).border = border

            ws.cell(row=cont, column=3).value = fac.proveedor.nombre
            ws.cell(row=cont, column=3).border = border

            ws.cell(row=cont, column=4).value = fac.total
            ws.cell(row=cont, column=4).border = border

            ws.cell(row=cont, column=5).value = fac.subTotal
            ws.cell(row=cont, column=5).border = border

            ws.cell(row=cont, column=6).value = fac.iva
            ws.cell(row=cont, column=6).border = border

            ws.cell(row=cont, column=7).value = fac.tipoCambio
            ws.cell(row=cont, column=7).border = border

            ws.cell(row=cont, column=8).value = fac.TotalDolar()
            ws.cell(row=cont, column=8).border = border

            ws.cell(row=cont, column=9).value = fac.detalle
            ws.cell(row=cont, column=9).border = border

            ws.cell(row=cont, column=10).value = fac.fecha
            ws.cell(row=cont, column=10).border = border

            ws.cell(row=cont, column=11).value = fac.empresa
            ws.cell(row=cont, column=11).border = border

            ws.cell(row=cont, column=12).value = fac.autorizado
            ws.cell(row=cont, column=12).border = border

            ws.cell(row=cont, column=13).value = fac.tipoMant
            ws.cell(row=cont, column=13).border = border

            ws.cell(row=cont, column=14).value = fac.comentarios
            ws.cell(row=cont, column=14).border = border

            ws.cell(row=cont, column=15).value = fac.realizado
            ws.cell(row=cont, column=15).border = border

            ws.cell(row=cont, column=16).value = fac.fechaMant
            ws.cell(row=cont, column=16).border = border

            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 35
            ws.column_dimensions['J'].width = 20
            ws.column_dimensions['K'].width = 20
            ws.column_dimensions['L'].width = 20
            ws.column_dimensions['M'].width = 30
            ws.column_dimensions['N'].width = 50
            ws.column_dimensions['O'].width = 20
            ws.column_dimensions['P'].width = 20

            cont += 1

        nombre_archivo = 'Reporte Mantenimiento.xlsx'
        response = HttpResponse(content_type='application/ms-excel')
        content = 'attachment; filename = {0}'.format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response