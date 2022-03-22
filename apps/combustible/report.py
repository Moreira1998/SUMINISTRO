from django.views.generic import TemplateView
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Side, Border, PatternFill, Alignment
from openpyxl.styles import Font
from apps.combustible.models import Consumo


# ------------------------------------------------------------------------------------>
# view Report Consumo


class ReporteExcelConsumo(TemplateView):
    def get(self, request, *args, **kwargs):
        consumo = Consumo.objects.all()
        wb = Workbook()
        ws = wb.active

        thin = Side(border_style="thin", color="000000")  # Estilo de frontera, color
        border = Border(left=thin, right=thin, top=thin, bottom=thin)  # Posición fronteriza

        titulo = Font(size=14, bold=True, name='Microsoft Yahei', color="4d148c")
        titulo2 = Font(size=14, bold=True, name='Microsoft Yahei', color="ff6600")
        font = Font(size=8, bold=True, name='Microsoft Yahei', color="000000")

        fill = PatternFill(patternType="solid", start_color="FFCD00")
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws['B2'] = 'Reporte Rendimiento De Combustible'
        ws.merge_cells('B2:E2')
        ws['B2'].font = titulo

        ws['B3'] = 'Passus Velox'
        ws.merge_cells('B3:E3')
        ws['B3'].font = titulo2

        ws['B5'] = 'Vahiculo '
        ws['B5'].font = font
        ws['B5'].alignment = align
        ws['B5'].border = border

        ws['C5'] = 'Vahiculo '
        ws['C5'].font = font
        ws['C5'].alignment = align
        ws['C5'].border = border

        ws['D5'] = 'Factura '
        ws['D5'].font = font
        ws['D5'].alignment = align
        ws['D5'].border = border

        ws['E5'] = 'Monto '
        ws['E5'].font = font
        ws['E5'].alignment = align
        ws['E5'].border = border

        ws['F5'] = 'Litros '
        ws['F5'].font = font
        ws['F5'].alignment = align
        ws['F5'].border = border

        ws['G5'] = 'KM Inicio '
        ws['G5'].font = font
        ws['G5'].alignment = align
        ws['G5'].border = border

        ws['H5'] = 'KM Final '
        ws['H5'].font = font
        ws['H5'].alignment = align
        ws['H5'].border = border

        ws['I5'] = 'Rendimiento '
        ws['I5'].font = font
        ws['I5'].alignment = align
        ws['I5'].border = border

        cont = 6
        for consumos in consumo:
            if consumos.Rendimiento() < consumos.vehiculo.rendimiento:

                ws.cell(row=cont, column=2).value = consumos.fecha
                ws.cell(row=cont, column=2).border = border
                ws.cell(row=cont, column=2).fill = fill

                ws.cell(row=cont, column=3).value = consumos.vehiculo.placa
                ws.cell(row=cont, column=3).border = border
                ws.cell(row=cont, column=3).fill = fill

                ws.cell(row=cont, column=4).value = consumos.factura
                ws.cell(row=cont, column=4).border = border
                ws.cell(row=cont, column=4).fill = fill

                ws.cell(row=cont, column=5).value = consumos.monto
                ws.cell(row=cont, column=5).border = border
                ws.cell(row=cont, column=5).fill = fill

                ws.cell(row=cont, column=6).value = consumos.litros
                ws.cell(row=cont, column=6).border = border
                ws.cell(row=cont, column=6).fill = fill

                ws.cell(row=cont, column=7).value = consumos.km_inicio
                ws.cell(row=cont, column=7).border = border
                ws.cell(row=cont, column=7).fill = fill

                ws.cell(row=cont, column=8).value = consumos.km_fin
                ws.cell(row=cont, column=8).border = border
                ws.cell(row=cont, column=8).fill = fill

                ws.cell(row=cont, column=9).value = consumos.Rendimiento()
                ws.cell(row=cont, column=9).border = border
                ws.cell(row=cont, column=9).fill = fill

                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 15
                ws.column_dimensions['I'].width = 15

                cont += 1
            else:
                ws.cell(row=cont, column=2).value = consumos.fecha
                ws.cell(row=cont, column=2).border = border

                ws.cell(row=cont, column=3).value = consumos.vehiculo.placa
                ws.cell(row=cont, column=3).border = border

                ws.cell(row=cont, column=4).value = consumos.factura
                ws.cell(row=cont, column=4).border = border

                ws.cell(row=cont, column=5).value = consumos.monto
                ws.cell(row=cont, column=5).border = border

                ws.cell(row=cont, column=6).value = consumos.litros
                ws.cell(row=cont, column=6).border = border

                ws.cell(row=cont, column=7).value = consumos.km_inicio
                ws.cell(row=cont, column=7).border = border

                ws.cell(row=cont, column=8).value = consumos.km_fin
                ws.cell(row=cont, column=8).border = border

                ws.cell(row=cont, column=9).value = consumos.Rendimiento()
                ws.cell(row=cont, column=9).border = border

                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 15
                ws.column_dimensions['I'].width = 15

                cont += 1

        nombre_archivo = 'Reporte Combustible.xlsx'
        response = HttpResponse(content_type='application/ms-excel')
        content = 'attachment; filename = {0}'.format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
