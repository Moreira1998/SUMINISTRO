from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.views.generic.base import RedirectView, TemplateView, View
from django.urls import reverse_lazy

from apps.producto.models import Producto
from apps.solicitud.forms import SolicitudForm
from apps.solicitud.models import Solicitud

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.styles import Font
from django.http.response import HttpResponse


class SolicitudList(ListView):
    model = Solicitud
    template_name = 'solicitud/solicitud_list.html'
    context_object_name = 'estadoSolicitud_list'
    queryset = Solicitud.objects.all()


class EstadoSolicitudList(ListView):
    model = Solicitud
    template_name = 'solicitud/estado/estadoSolicitud_list.html'
    context_object_name = 'estadoSolicitud_list'
    queryset = Solicitud.objects.all()


class SolicitudUpdate(UpdateView):
    model = Solicitud
    form_class = SolicitudForm
    template_name = 'solicitud/solicitud_form.html'
    success_url = reverse_lazy('solicitud:solicitud_list')


class SolicitudDelete(DeleteView):
    model = Solicitud
    form_class = SolicitudForm
    template_name = 'solicitud/solicitud_delete.html'
    success_url = reverse_lazy('solicitud:solicitud_list')


class SolicitudNew(CreateView):
    model = Solicitud
    form_class = SolicitudForm
    template_name = 'solicitud/solicitud_form.html'
    success_url = reverse_lazy('solicitud:estadoSolicitud_list')


class DespacharSolicitud(RedirectView):
    """ Despacha una solicitud restando la cantidad al
    stock del producto.

    la vista redirige a la lista de productos """

    url = reverse_lazy('solicitud:solicitud_list')

    def get(self, request, *args, **kwargs):
        # obtenemos la solicitud desde el id
        solicitud = Solicitud.objects.get(id=self.kwargs['pk'])
        # obtenemos el id del producto desde la solicitud
        producto = Producto.objects.get(id=solicitud.producto.id)
        # obtenemos la cantidad solicitada
        cantidad = solicitud.cantidad
        # la restamos del stock
        producto.cantidad -= cantidad
        # Cambiamos el status
        solicitud.status = True
        # guardamos los cambiosde producto
        producto.save()
        # guardamos los cambio de solicitud
        solicitud.save()

        return super().get(request, *args, **kwargs)


class ReporteExcelSolicitud(TemplateView):
    def get(self, request, *args, **kwargs):
        solicitud = Solicitud.objects.all()
        wb = Workbook()
        ws = wb.active

        thin = Side(border_style="thin", color="000000")  # Estilo de frontera, color
        border = Border(left=thin, right=thin, top=thin, bottom=thin)  # Posición fronteriza

        titulo = Font(size=14, bold=True, name='Microsoft Yahei', color="4d148c")
        titulo2 = Font(size=14, bold=True, name='Microsoft Yahei', color="ff6600")
        font = Font(size=8, bold=True, name='Microsoft Yahei', color="000000")

        fill = PatternFill(patternType="solid", start_color="ff6600")
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws['B2'] = 'Reporte De Solicitudes'
        ws.merge_cells('B2:E2')
        ws['B2'].font = titulo

        ws['B3'] = 'Passus Velox'
        ws.merge_cells('B3:E3')
        ws['B3'].font = titulo2

        ws['B5'] = 'producto '
        ws['B5'].font = font
        ws['B5'].alignment = align
        ws['B5'].border = border

        ws['C5'] = 'Area '
        ws['C5'].font = font
        ws['C5'].alignment = align
        ws['C5'].border = border

        ws['D5'] = 'Cantidad '
        ws['D5'].font = font
        ws['D5'].alignment = align
        ws['D5'].border = border

        ws['E5'] = 'Fecha '
        ws['E5'].font = font
        ws['E5'].alignment = align
        ws['E5'].border = border

        cont = 6

        for solicitudes in solicitud:
            if solicitudes.status == False:
                ws.cell(row=cont, column=2).value = solicitudes.producto.nombre
                ws.cell(row=cont, column=2).border = border
                ws.cell(row=cont, column=2).fill = fill

                ws.cell(row=cont, column=3).value = solicitudes.area.nombre
                ws.cell(row=cont, column=3).border = border

                ws.cell(row=cont, column=4).value = solicitudes.cantidad
                ws.cell(row=cont, column=4).border = border

                ws.cell(row=cont, column=5).value = solicitudes.fecha
                ws.cell(row=cont, column=5).border = border

                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15

                cont += 1
            else:

                ws.cell(row=cont, column=2).value = solicitudes.producto.nombre
                ws.cell(row=cont, column=2).border = border

                ws.cell(row=cont, column=3).value = solicitudes.area.nombre
                ws.cell(row=cont, column=3).border = border

                ws.cell(row=cont, column=4).value = solicitudes.cantidad
                ws.cell(row=cont, column=4).border = border

                ws.cell(row=cont, column=5).value = solicitudes.fecha
                ws.cell(row=cont, column=5).border = border

                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15

                cont += 1
        nombre_archivo = 'Reporte solicitudes.xlsx'
        response = HttpResponse(content_type='application/ms-excel')
        content = 'attachment; filename = {0}'.format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteExcelAprobados(TemplateView):
    def get(self, request, *args, **kwargs):
        solicitud = Solicitud.objects.all()
        wb = Workbook()
        ws = wb.active

        thin = Side(border_style="thin", color="000000")  # Estilo de frontera, color
        border = Border(left=thin, right=thin, top=thin, bottom=thin)  # Posición fronteriza

        titulo = Font(size=14, bold=True, name='Microsoft Yahei', color="4d148c")
        titulo2 = Font(size=14, bold=True, name='Microsoft Yahei', color="ff6600")
        font = Font(size=8, bold=True, name='Microsoft Yahei', color="000000")

        fill = PatternFill(patternType="solid", start_color="ff6600")
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws['B2'] = 'Reporte De Aprobados'
        ws.merge_cells('B2:E2')
        ws['B2'].font = titulo

        ws['B3'] = 'Passus Velox'
        ws.merge_cells('B3:E3')
        ws['B3'].font = titulo2

        ws['B5'] = 'Solicitado por: _________________________________________________'
        ws.merge_cells('B5:E5')

        ws['B7'] = 'Autorizado por: _________________________________________________'
        ws.merge_cells('B7:E7')

        ws['B9'] = 'producto '
        ws['B9'].font = font
        ws['B9'].alignment = align
        ws['B9'].border = border

        ws['C9'] = 'Area '
        ws['C9'].font = font
        ws['C9'].alignment = align
        ws['C9'].border = border

        ws['D9'] = 'Cantidad '
        ws['D9'].font = font
        ws['D9'].alignment = align
        ws['D9'].border = border

        ws['E9'] = 'Fecha '
        ws['E9'].font = font
        ws['E9'].alignment = align
        ws['E9'].border = border

        cont = 10

        for solicitudes in solicitud:
            if solicitudes.status == False:
                ws.cell(row=cont, column=2).value = solicitudes.producto.nombre
                ws.cell(row=cont, column=2).border = border

                ws.cell(row=cont, column=3).value = solicitudes.area.nombre
                ws.cell(row=cont, column=3).border = border

                ws.cell(row=cont, column=4).value = solicitudes.cantidad
                ws.cell(row=cont, column=4).border = border

                ws.cell(row=cont, column=5).value = solicitudes.fecha
                ws.cell(row=cont, column=5).border = border

                ws.column_dimensions['B'].width = 25
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 25
                ws.column_dimensions['E'].width = 25
                cont += 1
        nombre_archivo = 'Reporte solicitudes.xlsx'
        response = HttpResponse(content_type='application/ms-excel')
        content = 'attachment; filename = {0}'.format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteFiltrar(ListView):
    model = Solicitud
    template_name = 'solicitud/estado/reporte_list.html'
    context_object_name = 'reporte_list'

    def get_queryset(self):

        fecha_1 = self.request.GET.get('Fecha1', '')
        fecha_2 = self.request.GET.get('Fecha2', '')

        if fecha_1 and fecha_2:
            return Solicitud.objects.filter(fecha__range=(fecha_1, fecha_2))
        else:
            return Solicitud.objects.all()


class ReporteExcelProductos(TemplateView):
    def get(self, request, *args, **kwargs):
        producto = Producto.objects.all()
        wb = Workbook()
        ws = wb.active

        thin = Side(border_style="thin", color="000000")  # Estilo de frontera, color
        border = Border(left=thin, right=thin, top=thin, bottom=thin)  # Posición fronteriza

        titulo = Font(size=14, bold=True, name='Microsoft Yahei', color="4d148c")
        titulo2 = Font(size=14, bold=True, name='Microsoft Yahei', color="ff6600")
        font = Font(size=8, bold=True, name='Microsoft Yahei', color="000000")

        fill = PatternFill(patternType="solid", start_color="ff6600")
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws['B2'] = 'Reporte De Productos'
        ws.merge_cells('B2:E2')
        ws['B2'].font = titulo

        ws['B3'] = 'Passus Velox'
        ws.merge_cells('B3:E3')
        ws['B3'].font = titulo2

        ws['B5'] = 'ID '
        ws['B5'].font = font
        ws['B5'].alignment = align
        ws['B5'].border = border

        ws['C5'] = 'Nombre '
        ws['C5'].font = font
        ws['C5'].alignment = align
        ws['C5'].border = border

        ws['D5'] = 'Marca '
        ws['D5'].font = font
        ws['D5'].alignment = align
        ws['D5'].border = border

        ws['E5'] = 'Categoria '
        ws['E5'].font = font
        ws['E5'].alignment = align
        ws['E5'].border = border

        ws['F5'] = 'Cantidad '
        ws['F5'].font = font
        ws['F5'].alignment = align
        ws['F5'].border = border

        cont = 6

        for productos in producto:

            ws.cell(row=cont, column=2).value = productos.id
            ws.cell(row=cont, column=2).border = border

            ws.cell(row=cont, column=3).value = productos.nombre
            ws.cell(row=cont, column=3).border = border

            ws.cell(row=cont, column=4).value = productos.marca.nombre
            ws.cell(row=cont, column=4).border = border

            ws.cell(row=cont, column=5).value = productos.categoria.nombre
            ws.cell(row=cont, column=5).border = border

            ws.cell(row=cont, column=6).value = productos.cantidad
            ws.cell(row=cont, column=6).border = border

            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15

            cont += 1

        nombre_archivo = 'Reporte productos.xlsx'
        response = HttpResponse(content_type='application/ms-excel')
        content = 'attachment; filename = {0}'.format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response